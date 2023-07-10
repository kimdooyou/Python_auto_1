import openpyxl
import os

#새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 현재 활성화된 시트 선택
ws = wb.active

# 시트 이름 변경
ws.title = '자동화로만든겅미'

# 엑셀 저장
wb.save('자동화된엑셀.xlsx')


# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 새로운 시트 생성
wb.create_sheet('2024.04')

# 모든 시트 이름 출력
print(wb.sheetnames)

# 'Sheet'라는 시트가 있는지 확인하고 있으면 삭제
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 디렉토리가 존재하는지 확인하고 없다면 생성
dir_path = '04. 엑셀자동화'
if not os.path.exists(dir_path):
    os.makedirs(dir_path)

# 엑셀 저장
wb.save(os.path.join(dir_path, '화학물질관리대장.xlsx'))

save_path = '04. 엑셀자동화/화학물질관리대장.xlsx'

# 기존 엑셀 파일 불러오기
wb = openpyxl.load_workbook(save_path)

# 활성화된 시트 선택
ws = wb.active  

# 데이터 추가 (1)
ws['A1'] = '공정명'
ws['B1'] = '화학물질명'
ws['C1'] = '제조/사용'
ws['D1'] = '월취급량'
ws['E1'] = '단위'
ws['F1'] = '비고'
ws['G1'] = '법적규제'
ws['H1'] = 'CAS No.'
ws['I1'] = '규제현황'
ws['J1'] = '구성성분'

# 데이터 추가 (2)
ws.cell(row=2, column=1, value='공정A')
ws.cell(row=2, column=2, value='CR13')
ws.cell(row=2, column=3, value='사용')
ws.cell(row=2, column=4, value='100')
ws.cell(row=2, column=5, value='kg')
ws.cell(row=2, column=6, value='')
ws.cell(row=2, column=7, value='이산화티타늄')
ws.cell(row=2, column=8, value='13463-67-7')
ws.cell(row=2, column=9, value='규제중')
ws.cell(row=2, column=10, value='10~15%')


# 엑셀 저장
wb.save(save_path)

import os

# 새로운 엑셀 파일 생성
wb = openpyxl.Workbook()

# 새로운 시트 생성
wb.create_sheet('2024.04')

# 모든 시트 이름 출력
print(wb.sheetnames)

# 'Sheet'라는 시트가 있는지 확인하고 있으면 삭제
if 'Sheet' in wb.sheetnames:
    del wb['Sheet']

# 디렉토리가 존재하는지 확인하고 없다면 생성
dir_path = '03. 엑셀자동화'
if not os.path.exists(dir_path):
    os.makedirs(dir_path)

# 엑셀 저장
wb.save(os.path.join(dir_path, '화학물질관리대장.xlsx'))